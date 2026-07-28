from __future__ import annotations

import os
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT = Path(os.getenv("REPORT_OUTPUT", "outputs/adx_basic_report_20260725/7.25广告投放基础分析样例.xlsx"))
DAY = os.getenv("REPORT_DAY", "2026-07-25")
PREV_DAY = os.getenv(
    "REPORT_PREV_DAY",
    (datetime.strptime(DAY, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d"),
)
DAY_DT = datetime.strptime(DAY, "%Y-%m-%d")
DAY_LABEL = f"{DAY} 00:00—{DAY} 23:59"
DAY_SHORT = f"{DAY_DT.month}.{DAY_DT.day}"
ORDER_ID_TEXT = os.getenv("REPORT_ORDER_ID", "").strip()
ORDER_ID = int(ORDER_ID_TEXT) if ORDER_ID_TEXT else None


BLUE_DARK = "073D63"
BLUE = "2F70C9"
BLUE_LIGHT = "DDEBF7"
GREEN = "7FD0AA"
YELLOW = "FFE699"
RED = "FF0000"
GRID = "D9E2EC"
TEXT = "17324D"


def border() -> Border:
    side = Side(style="thin", color=GRID)
    return Border(left=side, right=side, top=side, bottom=side)


def style_range(ws, cell_range: str, fill: str | None = None, font_color: str = TEXT, bold: bool = False, size: int = 11):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Arial", size=size, bold=bold, color=font_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border()


def rate(v: float, d: float) -> float:
    return v / d if d else 0


def pct(v: float) -> str:
    return f"{v:.2%}"


def delta(v: float, p: float) -> str:
    v = v or 0
    p = p or 0
    if not p:
        return "前日为0"
    sign = "+" if v >= p else ""
    return f"{sign}{(v - p) / p:.2%}"


def day_key(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if len(text) >= 10 and text[:4].isdigit():
        return text[:10]
    return text


def find_header(ws, header: str) -> int | None:
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == header:
            return col
    return None


def remove_cpc_columns(wb):
    for ws in wb.worksheets:
        cols = []
        for row in range(1, min(ws.max_row, 20) + 1):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row, col).value
                if isinstance(val, str) and "CPC" in val:
                    cols.append(col)
        for col in sorted(set(cols), reverse=True):
            ws.delete_cols(col)


def restyle_dashboard(wb):
    source = wb["过滤情况"]
    header_row = 1
    for r in range(1, min(source.max_row, 20) + 1):
        vals = [source.cell(r, c).value for c in range(1, source.max_column + 1)]
        if "日期" in vals and "邀约" in vals:
            header_row = r
            break
    headers = [source.cell(header_row, c).value for c in range(1, source.max_column + 1)]
    rows = []
    for r in range(header_row + 1, source.max_row + 1):
        item = {headers[c - 1]: source.cell(r, c).value for c in range(1, source.max_column + 1)}
        if item.get("日期") is None:
            break
        rows.append(item)

    def by_day(day: str) -> dict:
        for item in rows:
            if day_key(item.get("日期")) == day:
                return {k: (0 if v is None else v) for k, v in item.items()}
        return {"日期": day, "邀约": 0, "出价": 0, "竞得": 0, "点击": 0, "落地页曝光": 0, "跳转": 0, "花销(元)": 0}

    prev = by_day(PREV_DAY)
    cur = by_day(DAY)

    ws = wb["大盘指标汇总"]
    ws.delete_rows(1, ws.max_row)
    ws.merge_cells("A1:H1")
    ws["A1"] = "广告投放情况概览"
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = f"时间：{DAY_LABEL}｜默认白名单过滤；无落地页/直接跳转订单漏斗到点击｜数据源：Doris ads.ods"
    ws["A2"].fill = PatternFill("solid", fgColor="EAF2FB")
    ws["A2"].font = Font(name="Arial", size=10, color="5D7285")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A4:H4")
    ws["A4"] = "总漏斗"
    style_range(ws, "A4:H4", BLUE_LIGHT, bold=True)
    metrics = ["总邀约", "出价", "竞得", "点击", "落地页曝光", "跳转", "总花销(元)"]
    values = [cur["邀约"], cur["出价"], cur["竞得"], cur["点击"], cur["落地页曝光"], cur["跳转"], cur["花销(元)"]]
    for i, m in enumerate(metrics, 1):
        ws.cell(5, i).value = m
        ws.cell(6, i).value = values[i - 1]
    style_range(ws, "A5:G5", BLUE, "FFFFFF", True)
    style_range(ws, "A6:G6", None, BLUE_DARK, True)
    ws["A7"] = f"过滤前：{(cur['邀约'] or 0) / 100000000:.2f}亿；相较前日出价 {delta(cur['出价'], prev['出价'])}，竞得 {delta(cur['竞得'], prev['竞得'])}，点击 {delta(cur['点击'], prev['点击'])}"
    ws.merge_cells("A7:H7")
    ws["A7"].font = Font(name="Arial", size=10, color=TEXT)
    ws["A7"].alignment = Alignment(horizontal="left")

    rate_headers = ["出价/邀约", "出价→竞得", "竞得→点击", "点击→落地页曝光", "落地页曝光→跳转", "总花销(元)"]
    rate_vals = [
        rate(cur["出价"], cur["邀约"]),
        rate(cur["竞得"], cur["出价"]),
        rate(cur["点击"], cur["竞得"]),
        rate(cur["落地页曝光"], cur["点击"]),
        rate(cur["跳转"], cur["落地页曝光"]),
        cur["花销(元)"],
    ]
    for i, m in enumerate(rate_headers, 1):
        ws.cell(8, i).value = m
        ws.cell(9, i).value = rate_vals[i - 1]
    style_range(ws, "A8:F8", BLUE, "FFFFFF", True)
    style_range(ws, "A9:F9", None, BLUE_DARK, True)
    ws["D9"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["E9"].font = Font(name="Arial", size=11, bold=True, color=RED)

    ws.merge_cells("A10:H10")
    ws["A10"] = f"相较前一天（{PREV_DAY}）"
    style_range(ws, "A10:H10", BLUE_LIGHT, bold=True)
    compare_headers = ["邀约", "出价", "竞得", "点击", "落地页曝光", "跳转", "总花销(元)"]
    for i, m in enumerate(compare_headers, 1):
        ws.cell(11, i).value = m
        ws.cell(12, i).value = cur[m.replace("总花销(元)", "花销(元)")] if m == "总花销(元)" else cur[m]
        pv = prev[m.replace("总花销(元)", "花销(元)")] if m == "总花销(元)" else prev[m]
        cv = ws.cell(12, i).value
        ws.cell(13, i).value = "0 → 0" if m == "跳转" and not cv and not pv else delta(cv, pv)
    style_range(ws, "A11:G11", GREEN, "FFFFFF", True)
    style_range(ws, "A12:G13", None, BLUE_DARK, True)

    ws.merge_cells("A16:H16")
    ws["A16"] = "初步结论 / 问题"
    style_range(ws, "A16:H16", BLUE_LIGHT, bold=True)
    if cur["落地页曝光"] > cur["点击"]:
        landing_note = f"pageExp={cur['落地页曝光']:,.0f} 高于 click={cur['点击']:,.0f}，需继续核对埋点关联或是否存在非点击来源访问。"
    else:
        landing_note = f"pageExp={cur['落地页曝光']:,.0f}，click={cur['点击']:,.0f}，落地页曝光已按点击后链路归因；无落地页订单不进入落地页漏斗。"
    conclusions = [
        ("1", "大盘", f"{DAY_SHORT} 出价 {cur['出价']:,.0f}，较前日 {delta(cur['出价'], prev['出价'])}；竞得 {cur['竞得']:,.0f}，整体投放强度显著放大。"),
        ("2", "漏斗", f"出价→竞得 {pct(rate(cur['竞得'], cur['出价']))}，竞得→点击 {pct(rate(cur['点击'], cur['竞得']))}；跳转仅 {cur['跳转']:,.0f}，落地页后链路偏弱。"),
        ("3", "落地页", landing_note),
        ("4", "订单", "直接跳转/无落地页订单漏斗到点击即可；有落地页订单继续看落地页曝光、完读、跳转链路。"),
        ("5", "花销", f"总花销 {cur['花销(元)']:,.2f} 元；后续花销专门分析可继续拆 WPR、APP、策略和广告位。"),
    ]
    ws.cell(17, 1).value = "序号"
    ws.cell(17, 2).value = "类型"
    ws.cell(17, 3).value = "内容"
    ws.merge_cells("C17:H17")
    style_range(ws, "A17:H17", BLUE, "FFFFFF", True)
    for r, row in enumerate(conclusions, 18):
        ws.cell(r, 1).value = row[0]
        ws.cell(r, 2).value = row[1]
        ws.cell(r, 3).value = row[2]
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    style_range(ws, f"A18:H{17 + len(conclusions)}")
    for c in range(1, 8):
        ws.cell(9, c).number_format = "0.00%" if c <= 5 else "#,##0.00"
    for c in range(1, 8):
        ws.cell(12, c).number_format = "#,##0.00" if c == 7 else "#,##0"
    for c in range(1, 8):
        ws.cell(13, c).number_format = "@"
    for c in range(1, 7):
        ws.cell(6, c).number_format = "#,##0"
    ws.cell(6, 7).number_format = "#,##0.00"
    widths = [18, 20, 24, 20, 22, 16, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def restyle_gantt(wb):
    ws = wb["投放策略甘特图"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    data = [dict(zip(headers, row)) for row in rows[1:] if row and row[0] is not None]
    hour_cols = [h for h in headers if isinstance(h, int)]
    first_hour_col = 8

    ws.delete_rows(1, ws.max_row)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=first_hour_col + len(hour_cols) - 1)
    ws["A1"] = "投放策略甘特图"
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=first_hour_col + len(hour_cols) - 1)
    ws["A2"] = "蓝色单元格=该策略在对应活跃小时有出价；时间轴仅保留全局活跃小时"
    ws["A2"].fill = PatternFill("solid", fgColor="EAF2FB")
    ws["A2"].font = Font(name="Arial", size=10, color="5D7285")
    ws["A2"].alignment = Alignment(horizontal="left")

    fixed = ["订单ID", "策略ID", "广告名称", "开始时间", "最后一次出价时间", "活跃小时", "出价"]
    for c, h in enumerate(fixed, 1):
        ws.cell(4, c).value = h
    if hour_cols:
        ws.merge_cells(start_row=4, start_column=first_hour_col, end_row=4, end_column=first_hour_col + len(hour_cols) - 1)
        ws.cell(4, first_hour_col).value = f"{DAY_DT:%m-%d}"
    for i, h in enumerate(hour_cols):
        ws.cell(5, first_hour_col + i).value = f"{h:02d}"
    style_range(ws, "A4:G5", BLUE, "FFFFFF", True)
    if hour_cols:
        style_range(ws, f"{get_column_letter(first_hour_col)}4:{get_column_letter(first_hour_col + len(hour_cols) - 1)}5", BLUE, "FFFFFF", True)

    for r, item in enumerate(data, 6):
        vals = [
            item.get("order_id"),
            item.get("strategy_id"),
            item.get("strategy_name") or item.get("order_name"),
            item.get("开始时间"),
            item.get("最后出价时间"),
            item.get("活跃小时"),
            item.get("出价"),
        ]
        for c, val in enumerate(vals, 1):
            ws.cell(r, c).value = val
            ws.cell(r, c).alignment = Alignment(horizontal="center" if c != 3 else "left", vertical="center")
            ws.cell(r, c).border = border()
        for i, h in enumerate(hour_cols):
            cell = ws.cell(r, first_hour_col + i)
            val = item.get(h) or 0
            cell.value = ""
            cell.border = border()
            if val:
                cell.fill = PatternFill("solid", fgColor="40A2D8")
        if ORDER_ID is not None and item.get("order_id") == ORDER_ID:
            for c in range(1, 8):
                ws.cell(r, c).fill = PatternFill("solid", fgColor="FFF2CC")

    widths = [12, 12, 34, 22, 22, 12, 14] + [6] * len(hour_cols)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 26 if row >= 4 else ws.row_dimensions[row].height
    ws.freeze_panes = "H6"
    ws.sheet_view.showGridLines = False


def sheet_subtitle(name: str) -> str:
    subtitles = {
        "过滤情况": f"以天为单位，对比 {DAY} 与前一天 {PREV_DAY}；前日对比用于替代样例中的4、5月份对比",
        "订单策略创意": "按出价量降序；同一 sheet 内依次为订单汇总、策略汇总、创意/素材汇总",
        "APP漏斗Top30": "按出价量 Top30 展示，并保留邀约、出价、竞得、点击、落地页曝光、跳转和比率",
        "落地页情况": "pageExp 作为落地页曝光；jumpButton 作为跳转；无落地页/直接跳转订单漏斗到点击即可",
        "小时分析": "按 request_time 小时归因；展示单日 24 小时漏斗、比率和花销",
        "APP×时段": "APP 与小时交叉分析；用于定位集中投放时段和异常峰值",
        "APP×订单策略素材": "APP 与订单策略素材交叉分析；用于定位 APP 与素材/策略适配关系",
        "时段×订单策略素材": "小时与订单策略素材交叉分析；用于定位策略投放窗口和小时效率",
        "口径说明": "数据源、漏斗定义、关联键和需确认口径",
    }
    return subtitles.get(name, "基础分析子表")


def is_header_row(ws, row: int, max_col: int | None = None) -> bool:
    max_col = max_col or ws.max_column
    vals = [ws.cell(row, c).value for c in range(1, min(max_col, 12) + 1)]
    labels = {str(v) for v in vals if v is not None}
    hits = {"日期", "指标", "APP名称", "app_bundle", "hour", "落地页ID", "order_id", "订单策略素材", "项目"} & labels
    return bool(hits) or ("出价" in labels and ("竞得" in labels or "点击" in labels))


def restyle_standard_sheet(ws):
    if ws.title in {"大盘指标汇总", "投放策略甘特图"}:
        return
    # Avoid stacking title rows if this script is run repeatedly.
    if ws["A1"].value != ws.title:
        ws.insert_rows(1, 3)
        last_col = max(8, ws.max_column)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        ws["A1"] = ws.title
        ws["A2"] = sheet_subtitle(ws.title)
    if ws.title == "小时分析":
        remove_zero_bid_hours(ws)
    move_spend_to_end(ws)
    last_col = ws.max_column
    last_row = ws.max_row
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].fill = PatternFill("solid", fgColor="EAF2FB")
    ws["A2"].font = Font(name="Arial", size=10, color="5D7285")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 24

    header_rows = [row for row in range(4, last_row + 1) if is_header_row(ws, row, last_col)]
    previous_header_for_row = {}
    current_header = None
    header_set = set(header_rows)
    for row in range(4, last_row + 1):
        if row in header_set:
            current_header = row
        previous_header_for_row[row] = current_header

    for row in range(4, last_row + 1):
        if row in header_set:
            for col in range(1, last_col + 1):
                cell = ws.cell(row, col)
                cell.fill = PatternFill("solid", fgColor=BLUE)
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border()
            ws.row_dimensions[row].height = 26
        else:
            empty = all(ws.cell(row, c).value is None for c in range(1, min(last_col, 12) + 1))
            for col in range(1, last_col + 1):
                cell = ws.cell(row, col)
                if empty:
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")
                    continue
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                cell.font = Font(name="Arial", size=10, color=TEXT)
                header_row = previous_header_for_row.get(row)
                header = str(ws.cell(header_row, col).value or "") if header_row else ""
                if isinstance(cell.value, float):
                    if "率" in header or "→" in header or "变化率" in header:
                        cell.number_format = "0.00%"
                    elif "花销" in header:
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "#,##0.00"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
                if "跳转" in header and isinstance(cell.value, (int, float)) and cell.value <= 5 and cell.value != 0:
                    cell.font = Font(name="Arial", size=10, bold=True, color=RED)
                if "落地页曝光" in header or "点击→落地页曝光" in header:
                    if isinstance(cell.value, float) and cell.value > 1:
                        cell.fill = PatternFill("solid", fgColor=YELLOW)
                cell.alignment = Alignment(horizontal="right" if isinstance(cell.value, (int, float)) else "left", vertical="center", wrap_text=True)
                cell.border = border()

    for col in range(1, last_col + 1):
        max_len = 0
        for row in range(1, min(last_row, 250) + 1):
            val = ws.cell(row, col).value
            max_len = max(max_len, min(len(str(val)) if val is not None else 0, 50))
        ws.column_dimensions[get_column_letter(col)].width = max(10, min(max_len + 2, 36))
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def remove_zero_bid_hours(ws):
    header_row = 4
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    if "出价" not in headers:
        return
    bid_col = headers.index("出价") + 1
    for row in range(ws.max_row, header_row, -1):
        hour = ws.cell(row, 1).value
        bid = ws.cell(row, bid_col).value
        if isinstance(hour, int) and (bid is None or bid == 0):
            ws.delete_rows(row, 1)


def move_spend_to_end(ws):
    header_row = 4
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    if "花销(元)" not in headers:
        return
    spend_col = headers.index("花销(元)") + 1
    non_empty_cols = [c for c, v in enumerate(headers, 1) if v is not None]
    if not non_empty_cols:
        return
    target_col = max(non_empty_cols)
    if spend_col >= target_col:
        return
    values = [ws.cell(r, spend_col).value for r in range(1, ws.max_row + 1)]
    styles = []
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, spend_col)
        styles.append((cell.font.copy(), cell.fill.copy(), cell.border.copy(), cell.alignment.copy(), cell.number_format))
    ws.delete_cols(spend_col)
    if spend_col < target_col:
        target_col -= 1
    ws.insert_cols(target_col + 1)
    for r, value in enumerate(values, 1):
        cell = ws.cell(r, target_col + 1)
        cell.value = value
        font, fill, bdr, align, fmt = styles[r - 1]
        cell.font = font
        cell.fill = fill
        cell.border = bdr
        cell.alignment = align
        cell.number_format = fmt


def main():
    wb = load_workbook(REPORT)
    remove_cpc_columns(wb)
    restyle_dashboard(wb)
    restyle_gantt(wb)
    for ws in wb.worksheets:
        restyle_standard_sheet(ws)
    wb.save(REPORT)
    print(REPORT.resolve())


if __name__ == "__main__":
    main()
