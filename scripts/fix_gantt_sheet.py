from __future__ import annotations

import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".vendor_pymysql"))
import pymysql  # noqa: E402


REPORT = Path(os.getenv("REPORT_OUTPUT", "outputs/adx_basic_report_20260725/7.25广告投放基础分析样例.xlsx"))
DAY = os.getenv("REPORT_DAY", "2026-07-25")
DAY_DT = datetime.strptime(DAY, "%Y-%m-%d")
ORDER_ID_TEXT = os.getenv("REPORT_ORDER_ID", "").strip()
ORDER_ID = int(ORDER_ID_TEXT) if ORDER_ID_TEXT else None
BLUE_DARK = "073D63"
BLUE = "2F70C9"
BLUE_CELL = "40A2D8"
GRID = "D9E2EC"
TEXT = "17324D"


def conn() -> pymysql.Connection:
    return pymysql.connect(
        host=os.getenv("DORIS_HOST", "127.0.0.1"),
        port=int(os.getenv("DORIS_PORT", "19030")),
        user=os.getenv("DORIS_USER", "WishFox"),
        password=os.environ["DORIS_PASSWORD"],
        database=os.getenv("DORIS_DATABASE", "ads"),
        connect_timeout=10,
        read_timeout=int(os.getenv("DORIS_READ_TIMEOUT", "900")),
        write_timeout=int(os.getenv("DORIS_WRITE_TIMEOUT", "900")),
        charset="utf8mb4",
    )


def border() -> Border:
    side = Side(style="thin", color=GRID)
    return Border(left=side, right=side, top=side, bottom=side)


def order_filter() -> str:
    return f" and order_id = {ORDER_ID}" if ORDER_ID is not None else ""


def main() -> None:
    sql = f"""
    with
    hourly as (
      select order_id, strategy_id, hour(response_time) h, count(*) hour_bid
      from ods.ods_ad_adx_bid_response_rt
      where response_time >= %s and response_time < date_add(%s, interval 1 day)
      {order_filter()}
      group by order_id, strategy_id, hour(response_time)
    ),
    total as (
      select order_id, strategy_id,
        min(response_time) start_time,
        max(response_time) last_time,
        count(distinct hour(response_time)) active_hours,
        count(*) bid_count
      from ods.ods_ad_adx_bid_response_rt
      where response_time >= %s and response_time < date_add(%s, interval 1 day)
      {order_filter()}
      group by order_id, strategy_id
    ),
    dim_strategy as (
      select strategy_id, max(strategy_name) strategy_name
      from dim.dim_ad_dsp_strategy
      group by strategy_id
    )
    select
      h.order_id,
      h.strategy_id,
      coalesce(d.strategy_name, cast(h.strategy_id as string)) ad_name,
      t.start_time,
      t.last_time,
      t.active_hours,
      t.bid_count,
      h.h,
      h.hour_bid
    from hourly h
    join total t on h.order_id=t.order_id and h.strategy_id=t.strategy_id
    left join dim_strategy d on h.strategy_id=d.strategy_id
    order by t.bid_count desc, h.order_id, h.strategy_id, h.h
    """
    attempts = int(os.getenv("DORIS_QUERY_RETRIES", "3"))
    for attempt in range(1, attempts + 1):
        try:
            with conn() as c:
                df = pd.read_sql(sql, c, params=(DAY, DAY, DAY, DAY))
            break
        except pymysql.err.OperationalError as exc:
            code = exc.args[0] if exc.args else None
            if code not in {2006, 2013} or attempt >= attempts:
                raise
            wait_seconds = min(2**attempt, 8)
            print(f"Doris query lost connection ({code}); retry {attempt}/{attempts - 1} after {wait_seconds}s", file=sys.stderr)
            time.sleep(wait_seconds)

    rows = []
    for (order_id, strategy_id, ad_name), g in df.groupby(["order_id", "strategy_id", "ad_name"], dropna=False):
        first = g.iloc[0]
        item = {
            "订单ID": order_id,
            "策略ID": strategy_id,
            "广告名称": ad_name,
            "开始时间": g["start_time"].min(),
            "最后一次出价时间": g["last_time"].max(),
            "活跃小时": int(g["active_hours"].max()),
            "出价": int(g["bid_count"].max()),
        }
        for _, r in g.iterrows():
            item[int(r["h"])] = int(r["hour_bid"])
        rows.append(item)
    rows.sort(key=lambda x: x["出价"], reverse=True)
    hours = sorted({int(h) for h in df["h"].dropna().unique()})

    wb = load_workbook(REPORT)
    if "投放策略甘特图" in wb.sheetnames:
        idx = wb.sheetnames.index("投放策略甘特图")
        del wb["投放策略甘特图"]
        ws = wb.create_sheet("投放策略甘特图", idx)
    else:
        ws = wb.create_sheet("投放策略甘特图")

    total_cols = 7 + len(hours)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = "投放策略甘特图"
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws["A2"] = "蓝色单元格=该策略在对应活跃小时有出价；单元格数字=该策略该小时出价数；时间轴仅保留全局活跃小时"
    ws["A2"].fill = PatternFill("solid", fgColor="EAF2FB")
    ws["A2"].font = Font(name="Arial", size=10, color="5D7285")
    ws["A2"].alignment = Alignment(horizontal="left")

    fixed = ["订单ID", "策略ID", "广告名称", "开始时间", "最后一次出价时间", "活跃小时", "出价"]
    for c, h in enumerate(fixed, 1):
        ws.cell(4, c).value = h
        ws.merge_cells(start_row=4, start_column=c, end_row=5, end_column=c)
    ws.merge_cells(start_row=4, start_column=8, end_row=4, end_column=total_cols)
    ws.cell(4, 8).value = f"{DAY_DT:%m-%d}"
    for i, h in enumerate(hours, 8):
        ws.cell(5, i).value = f"{h:02d}"

    for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=total_cols):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border()

    for r, item in enumerate(rows, 6):
        for c, key in enumerate(fixed, 1):
            cell = ws.cell(r, c)
            cell.value = item[key]
            cell.font = Font(name="Arial", size=10, color=TEXT)
            cell.alignment = Alignment(horizontal="left" if key == "广告名称" else "center", vertical="center")
            cell.border = border()
            if key == "出价":
                cell.number_format = "#,##0"
        for c, h in enumerate(hours, 8):
            cell = ws.cell(r, c)
            val = item.get(h, 0)
            cell.value = val if val else None
            cell.border = border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if val:
                cell.fill = PatternFill("solid", fgColor=BLUE_CELL)
                cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                cell.number_format = "#,##0"

    widths = [12, 12, 34, 22, 22, 12, 14] + [7] * len(hours)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 24
    ws.freeze_panes = "H6"
    ws.sheet_view.showGridLines = False
    wb.save(REPORT)
    print(REPORT.resolve())
    print("gantt_rows", len(rows), "hours", hours)


if __name__ == "__main__":
    main()
