from __future__ import annotations

import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from scripts.build_725_spend_report import (  # noqa: E402
    BLUE,
    BLUE_CELL,
    BLUE_DARK,
    GRID,
    TEXT,
    conn,
    load_app_mapping,
    pct,
    q,
    style_workbook,
    write_sheet,
)


DAY = os.getenv("REPORT_DAY", "2026-07-25")
PREV_DAY = os.getenv(
    "REPORT_PREV_DAY",
    (datetime.strptime(DAY, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d"),
)
DAY_DT = datetime.strptime(DAY, "%Y-%m-%d")
OUT = Path(
    os.getenv(
        "REPORT_OUTPUT",
        f"outputs/adx_bidding_report_{DAY_DT:%Y%m%d}/{DAY_DT.month}.{DAY_DT.day}广告投放竞价分析报表.xlsx",
    )
)
ORDER_ID_TEXT = os.getenv("REPORT_ORDER_ID", "").strip()
ORDER_ID = int(ORDER_ID_TEXT) if ORDER_ID_TEXT else None


def order_filter(alias: str = "r") -> str:
    return f" and {alias}.order_id = {ORDER_ID}" if ORDER_ID is not None else ""


def request_order_exists() -> str:
    if ORDER_ID is None:
        return ""
    return f"""
        and exists (
          select 1
          from ods.ods_ad_adx_bid_response_rt r
          where r.response_id=request_id
            and r.bid_imp_id=imp_id
            and r.order_id={ORDER_ID}
            and r.response_time >= request_time
            and r.response_time < date_add(request_time, interval 1 day)
        )
    """


def price_bucket(v) -> str:
    if pd.isna(v):
        return "未知"
    v = float(v)
    if v <= 100:
        return "100"
    if v <= 500:
        return "100-500"
    if v <= 1000:
        return "500-1000"
    if v <= 1500:
        return "1000-1500"
    if v <= 2000:
        return "1500-2000"
    return "2000+"


def add_bidding_rates(df: pd.DataFrame, total_bid: float | None = None, has_invite: bool = False) -> pd.DataFrame:
    df = df.copy()
    if has_invite and "邀约" in df:
        df["邀约→出价"] = pct(df["出价"], df["邀约"])
    df["出价→竞得"] = pct(df["竞得"], df["出价"])
    df["竞得→点击"] = pct(df["点击"], df["竞得"])
    df["点击→落地页曝光"] = pct(df["落地页曝光"], df["点击"])
    df["落地页曝光→跳转"] = pct(df["跳转"], df["落地页曝光"])
    if "花费(元)" in df:
        df["每次竞得花费"] = pct(df["花费(元)"], df["竞得"])
        df["每次点击花费"] = pct(df["花费(元)"], df["点击"])
    if total_bid is not None:
        df["出价占比"] = df["出价"] / total_bid if total_bid else 0
    return df


def request_agg(key_sql: str, label: str) -> pd.DataFrame:
    return q(
        f"""
        select {key_sql} {label}, count(*) 邀约
        from ods.ods_ad_adx_bid_request_rt
        where request_time >= %s and request_time < date_add(%s, interval 1 day)
        {request_order_exists()}
        group by {key_sql}
        """,
        (DAY, DAY),
    )


def format_compare_block(ws) -> None:
    count_metrics = {"邀约", "出价", "竞得", "点击", "落地页曝光", "跳转"}
    for row in range(1, ws.max_row + 1):
        labels = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 8) + 1)]
        if labels[:5] == ["指标", DAY, PREV_DAY, "较前日变化", "较前日变化率"]:
            for r in range(row + 1, ws.max_row + 1):
                metric = ws.cell(r, 1).value
                if metric is None:
                    break
                for c in (2, 3, 4):
                    cell = ws.cell(r, c)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0" if metric in count_metrics else "#,##0.00"
                if isinstance(ws.cell(r, 5).value, (int, float)):
                    ws.cell(r, 5).number_format = "0.00%"
            return


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    app_map = load_app_mapping()

    daily_sql = """
    with
    req as (
      select date(request_time) d, count(*) invite
      from ods.ods_ad_adx_bid_request_rt
      where request_time >= %s and request_time < date_add(%s, interval 2 day)
      group by date(request_time)
    ),
    bid as (
      select date(response_time) d, count(*) bid, count(distinct strategy_id) strategy_cnt,
             count(distinct hour(response_time)) active_hours, avg(bid_price) avg_bid_price,
             min(bid_price) min_bid_price, max(bid_price) max_bid_price
      from ods.ods_ad_adx_bid_response_rt
      where response_time >= %s and response_time < date_add(%s, interval 2 day)
      group by date(response_time)
    ),
    win as (
      select date(response_time) d, count(*) win, sum(coalesce(cast(nullif(wpr, '') as double),0))/100000 spend
      from ods.ods_ad_adx_bid_response_result_rt
      where response_time >= %s and response_time < date_add(%s, interval 2 day)
      group by date(response_time)
    ),
    clk as (
      select date(response_time) d, count(*) click
      from ods.ods_ad_adx_bid_response_click_rt
      where response_time >= %s and response_time < date_add(%s, interval 2 day)
      group by date(response_time)
    ),
    clk_pairs as (
      select date(response_time) d, response_id, bid_seat_id
      from ods.ods_ad_adx_bid_response_click_rt
      where response_time >= %s and response_time < date_add(%s, interval 2 day)
      group by date(response_time), response_id, bid_seat_id
    ),
    track_keys as (
      select
        coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]rid=([^&]+)', 1), ''), t.root_id) response_id_key,
        coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]pyck=([^&]+)', 1), ''), t.parent_event_id) bid_seat_id_key,
        t.event_code
      from ods.ods_track_event_default_raw t
      where t.event_time >= %s and t.event_time < date_add(%s, interval 2 day)
    ),
    trk as (
      select c.d,
        count(distinct case when t.event_code='pageExp' then concat(t.response_id_key,'#',t.bid_seat_id_key) end) page_exp,
        count(distinct case when t.event_code='jumpButton' then concat(t.response_id_key,'#',t.bid_seat_id_key) end) jump
      from clk_pairs c
      join track_keys t on t.response_id_key=c.response_id and t.bid_seat_id_key=c.bid_seat_id
      group by c.d
    )
    select req.d 日期, invite 邀约, coalesce(bid,0) 出价, coalesce(win,0) 竞得, coalesce(click,0) 点击,
           coalesce(page_exp,0) 落地页曝光, coalesce(jump,0) 跳转, coalesce(spend,0) `花费(元)`,
           coalesce(active_hours,0) 活跃小时, coalesce(strategy_cnt,0) 投放策略数,
           coalesce(avg_bid_price,0) 平均出价, coalesce(min_bid_price,0) 最低出价, coalesce(max_bid_price,0) 最高出价
    from req
    left join bid on req.d=bid.d
    left join win on req.d=win.d
    left join clk on req.d=clk.d
    left join trk on req.d=trk.d
    order by req.d
    """
    daily_params = (PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY)
    if ORDER_ID is not None:
        daily_sql = f"""
        with
        base as (
          select date(r.response_time) d, r.response_time, r.response_id, r.bid_seat_id, r.strategy_id, r.bid_price
          from ods.ods_ad_adx_bid_response_rt r
          where r.response_time >= %s and r.response_time < date_add(%s, interval 2 day)
            and r.order_id = {ORDER_ID}
        ),
        req as (
          select d, count(*) invite
          from base
          group by d
        ),
        bid as (
          select d, count(*) bid, count(distinct strategy_id) strategy_cnt,
                 count(distinct hour(response_time)) active_hours, avg(bid_price) avg_bid_price,
                 min(bid_price) min_bid_price, max(bid_price) max_bid_price
          from base
          group by d
        ),
        win as (
          select b.d, count(*) win, sum(coalesce(cast(nullif(w.wpr, '') as double),0))/100000 spend
          from base b
          join ods.ods_ad_adx_bid_response_result_rt w
            on b.response_id=w.response_id and b.bid_seat_id=w.bid_seat_id
          where w.response_time >= %s and w.response_time < date_add(%s, interval 2 day)
          group by b.d
        ),
        clk_pairs as (
          select b.d, c.response_id, c.bid_seat_id
          from base b
          join ods.ods_ad_adx_bid_response_click_rt c
            on b.response_id=c.response_id and b.bid_seat_id=c.bid_seat_id
          where c.response_time >= %s and c.response_time < date_add(%s, interval 2 day)
          group by b.d, c.response_id, c.bid_seat_id
        ),
        clk as (
          select d, count(*) click
          from clk_pairs
          group by d
        ),
        track_keys as (
          select
            coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]rid=([^&]+)', 1), ''), t.root_id) response_id_key,
            coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]pyck=([^&]+)', 1), ''), t.parent_event_id) bid_seat_id_key,
            t.event_code
          from ods.ods_track_event_default_raw t
          where t.event_time >= %s and t.event_time < date_add(%s, interval 2 day)
        ),
        trk as (
          select c.d,
            count(distinct case when t.event_code='pageExp' then concat(t.response_id_key,'#',t.bid_seat_id_key) end) page_exp,
            count(distinct case when t.event_code='jumpButton' then concat(t.response_id_key,'#',t.bid_seat_id_key) end) jump
          from clk_pairs c
          join track_keys t on t.response_id_key=c.response_id and t.bid_seat_id_key=c.bid_seat_id
          group by c.d
        )
        select req.d 日期, invite 邀约, coalesce(bid,0) 出价, coalesce(win,0) 竞得, coalesce(click,0) 点击,
               coalesce(page_exp,0) 落地页曝光, coalesce(jump,0) 跳转, coalesce(spend,0) `花费(元)`,
               coalesce(active_hours,0) 活跃小时, coalesce(strategy_cnt,0) 投放策略数,
               coalesce(avg_bid_price,0) 平均出价, coalesce(min_bid_price,0) 最低出价, coalesce(max_bid_price,0) 最高出价
        from req
        left join bid on req.d=bid.d
        left join win on req.d=win.d
        left join clk on req.d=clk.d
        left join trk on req.d=trk.d
        order by req.d
        """
        daily_params = (PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY)
    daily = q(daily_sql, daily_params)
    current = daily[daily["日期"].astype(str) == DAY].copy()
    prev = daily[daily["日期"].astype(str) == PREV_DAY].copy()
    if current.empty:
        raise RuntimeError(f"{DAY} 没有查询到数据")
    total_bid = float(current.iloc[0]["出价"])
    total_spend = float(current.iloc[0]["花费(元)"])

    wide_sql = f"""
    with
    win as (
      select response_id, bid_seat_id, count(*) win, sum(coalesce(cast(nullif(wpr, '') as double),0))/100000 spend_yuan
      from ods.ods_ad_adx_bid_response_result_rt
      where response_time >= %s and response_time < date_add(%s, interval 1 day)
      group by response_id, bid_seat_id
    ),
    clk as (
      select response_id, bid_seat_id, count(*) click
      from ods.ods_ad_adx_bid_response_click_rt
      where response_time >= %s and response_time < date_add(%s, interval 1 day)
      group by response_id, bid_seat_id
    ),
    track_keys as (
      select
        coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]rid=([^&]+)', 1), ''), t.root_id) response_id_key,
        coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]pyck=([^&]+)', 1), ''), t.parent_event_id) bid_seat_id_key,
        t.event_code
      from ods.ods_track_event_default_raw t
      where t.event_time >= %s and t.event_time < date_add(%s, interval 1 day)
    ),
    trk as (
      select t.response_id_key, t.bid_seat_id_key,
        max(case when t.event_code='pageExp' then 1 else 0 end) page_exp,
        max(case when t.event_code='jumpButton' then 1 else 0 end) jump
      from track_keys t
      join clk c on t.response_id_key=c.response_id and t.bid_seat_id_key=c.bid_seat_id
      group by t.response_id_key, t.bid_seat_id_key
    ),
    dim_strategy as (
      select strategy_id, strategy_name, order_id, order_name
      from (
        select strategy_id, strategy_name, order_id, order_name,
               row_number() over(partition by strategy_id order by etl_time desc) rn
        from dim.dim_ad_dsp_strategy
      ) s
      where rn=1
    )
    select
      r.response_time,
      q.request_time,
      hour(q.request_time) hour,
      coalesce(nullif(q.app_bundle,''),'(空)') app_bundle,
      r.response_id,
      r.bid_seat_id,
      r.bid_price,
      r.order_id,
      r.strategy_id,
      r.creative_id,
      coalesce(s.order_name, d.order_name, cast(r.order_id as string)) order_name,
      coalesce(s.strategy_name, d.strategy_name, cast(r.strategy_id as string)) strategy_name,
      coalesce(d.creative_name, cast(r.creative_id as string)) creative_name,
      d.material_id,
      d.material_type,
      coalesce(win.win,0) win,
      coalesce(win.spend_yuan,0) spend_yuan,
      coalesce(clk.click,0) click,
      coalesce(trk.page_exp,0) page_exp,
      coalesce(trk.jump,0) jump
    from ods.ods_ad_adx_bid_response_rt r
    left join ods.ods_ad_adx_bid_request_rt q
      on q.request_id=r.response_id and q.imp_id=r.bid_imp_id
      and q.request_time >= %s and q.request_time < date_add(%s, interval 1 day)
    left join win on r.response_id=win.response_id and r.bid_seat_id=win.bid_seat_id
    left join clk on r.response_id=clk.response_id and r.bid_seat_id=clk.bid_seat_id
    left join trk on r.response_id=trk.response_id_key and r.bid_seat_id=trk.bid_seat_id_key
    left join dim_strategy s on r.strategy_id=s.strategy_id
    left join dim.dim_ad_dsp_creative d on trim(r.creative_id)=cast(d.creative_id as string)
    where r.response_time >= %s and r.response_time < date_add(%s, interval 1 day)
    {order_filter("r")}
    """
    wide = q(wide_sql, (DAY, DAY, DAY, DAY, DAY, DAY, DAY, DAY, DAY, DAY))
    wide["APP名称"] = wide["app_bundle"].map(app_map).fillna(wide["app_bundle"])
    wide["出价分层"] = wide["bid_price"].map(price_bucket)
    wide["订单策略素材"] = (
        wide["order_name"].fillna("").astype(str)
        + " / "
        + wide["strategy_name"].fillna("").astype(str)
        + " / "
        + wide["creative_name"].fillna("").astype(str)
    )

    def bid_agg(keys: list[str]) -> pd.DataFrame:
        g = wide.groupby(keys, dropna=False).agg(
            出价=("response_id", "size"),
            竞得=("win", "sum"),
            点击=("click", "sum"),
            落地页曝光=("page_exp", "sum"),
            跳转=("jump", "sum"),
            花费=("spend_yuan", "sum"),
            平均出价=("bid_price", "mean"),
            最低出价=("bid_price", "min"),
            最高出价=("bid_price", "max"),
            活跃小时=("hour", "nunique"),
        ).reset_index()
        g = g.rename(columns={"花费": "花费(元)"})
        return add_bidding_rates(g, total_bid=total_bid)

    app_req = request_agg("coalesce(nullif(app_bundle,''),'(空)')", "app_bundle")
    app_req["APP名称"] = app_req["app_bundle"].map(app_map).fillna(app_req["app_bundle"])
    app = pd.merge(app_req, bid_agg(["APP名称", "app_bundle"]), on=["APP名称", "app_bundle"], how="right").fillna(0)
    app = add_bidding_rates(app, total_bid=total_bid, has_invite=True).sort_values("出价", ascending=False).head(30)
    app = app[["APP名称", "app_bundle"] + [c for c in app.columns if c not in {"APP名称", "app_bundle"}]]
    top_app_bundles = set(app["app_bundle"].astype(str))
    wide_top_app = wide[wide["app_bundle"].astype(str).isin(top_app_bundles)].copy()

    hour_req = request_agg("hour(request_time)", "hour")
    hour = pd.merge(hour_req, bid_agg(["hour"]), on="hour", how="right").fillna(0)
    hour = add_bidding_rates(hour, total_bid=total_bid, has_invite=True).sort_values("hour")

    order_material = bid_agg(["order_id", "strategy_id", "creative_id", "order_name", "strategy_name", "creative_name", "material_id", "material_type", "订单策略素材"]).sort_values("出价", ascending=False).head(300)
    bucket = bid_agg(["出价分层"]).sort_values("平均出价")

    def top_app_bid_agg(keys: list[str]) -> pd.DataFrame:
        g = wide_top_app.groupby(keys, dropna=False).agg(
            出价=("response_id", "size"),
            竞得=("win", "sum"),
            点击=("click", "sum"),
            落地页曝光=("page_exp", "sum"),
            跳转=("jump", "sum"),
            花费=("spend_yuan", "sum"),
            平均出价=("bid_price", "mean"),
            最低出价=("bid_price", "min"),
            最高出价=("bid_price", "max"),
            活跃小时=("hour", "nunique"),
        ).reset_index()
        g = g.rename(columns={"花费": "花费(元)"})
        return add_bidding_rates(g, total_bid=total_bid)

    crosses = {
        "APP×时段": top_app_bid_agg(["APP名称", "app_bundle", "hour"]).sort_values("出价", ascending=False).head(300),
        "APP×订单策略素材": top_app_bid_agg(["APP名称", "app_bundle", "订单策略素材"]).sort_values("出价", ascending=False).head(300),
        "APP×出价分层": top_app_bid_agg(["APP名称", "app_bundle", "出价分层"]).sort_values("出价", ascending=False).head(300),
        "时段×订单策略素材": bid_agg(["hour", "订单策略素材"]).sort_values(["hour", "出价"], ascending=[True, False]).head(300),
        "时段×出价分层": bid_agg(["hour", "出价分层"]).sort_values(["hour", "平均出价"]),
        "订单策略素材×出价分层": bid_agg(["订单策略素材", "出价分层"]).sort_values("出价", ascending=False).head(300),
    }

    dashboard = add_bidding_rates(current.copy(), total_bid=total_bid, has_invite=True)
    if not prev.empty:
        p = prev.iloc[0]
        c = current.iloc[0]
        compare = pd.DataFrame(
            [
                {"指标": m, DAY: c[m], PREV_DAY: p[m], "较前日变化": c[m] - p[m], "较前日变化率": (c[m] - p[m]) / p[m] if p[m] else 0}
                for m in ["邀约", "出价", "竞得", "点击", "落地页曝光", "跳转", "花费(元)", "平均出价"]
            ]
        )
    else:
        compare = pd.DataFrame()

    notes = pd.DataFrame(
        {
            "项目": ["竞价口径", "出价分层", "漏斗定义", "订单限定", "埋点关联", "策略名称"],
            "说明": [
                "竞价分析以 response 表出价为主，竞得按 result 表，点击按 click 表，花费=sum(wpr)/100000。",
                "出价分层按 bid_price：100、100-500、500-1000、1000-1500、1500-2000、2000+。",
                "邀约、出价、竞得、点击、落地页曝光(pageExp)、跳转(jumpButton)。",
                f"当前仅分析订单 {ORDER_ID}；请求表无订单字段，邀约统计为可关联到该订单出价的请求机会。" if ORDER_ID is not None else "未限定订单，按全量投放统计。",
                "track 优先从 preset_attributes.url 解析 rid/pyck 关联点击；解析不到时用 root_id/parent_event_id 兜底。",
                "strategy_id 名称优先来自 dim.dim_ad_dsp_strategy，创意/素材字段来自 dim.dim_ad_dsp_creative。",
            ],
        }
    )

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        write_sheet(
            writer,
            "大盘竞价概览",
            "广告投放竞价专项分析",
            f"时间：{DAY} 00:00—{DAY} 23:59｜按出价 response 维度分析｜出价分层基于 bid_price",
            dashboard,
        )
        if not compare.empty:
            compare.to_excel(writer, sheet_name="大盘竞价概览", index=False, startrow=len(dashboard) + 7)
            writer.book["大盘竞价概览"].cell(len(dashboard) + 6, 1).value = f"相较前一天（{PREV_DAY}）"

        wb = writer.book
        ws = wb.create_sheet("投放策略甘特图", 1)
        hour_cols = sorted(int(x) for x in wide["hour"].dropna().unique())
        total_cols = 8 + len(hour_cols)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws["A1"] = "投放策略出价甘特图"
        ws["A2"] = "蓝色单元格=该策略该小时有出价；数字=该小时出价数；仅展示单日活跃小时"
        fixed = ["订单ID", "策略ID", "策略名称", "开始时间", "结束时间", "活跃小时", "总出价", "平均出价"]
        for cidx, h in enumerate(fixed, 1):
            ws.cell(4, cidx).value = h
        for i, h in enumerate(hour_cols, 9):
            ws.cell(4, i).value = f"{h:02d}"
        ws.cell(5, 1).value = "逐小时总出价"
        ws.cell(5, 7).value = total_bid
        hourly_bid = wide.groupby("hour")["response_id"].size().to_dict()
        for i, h in enumerate(hour_cols, 9):
            ws.cell(5, i).value = int(hourly_bid.get(h, 0))
        gantt = wide.groupby(["order_id", "strategy_id", "strategy_name"], dropna=False).agg(
            开始时间=("response_time", "min"),
            结束时间=("response_time", "max"),
            活跃小时=("hour", "nunique"),
            总出价=("response_id", "size"),
            平均出价=("bid_price", "mean"),
        ).reset_index().sort_values("总出价", ascending=False)
        gantt_h = wide.groupby(["order_id", "strategy_id", "hour"], dropna=False).size().reset_index(name="小时出价")
        for ridx, row in enumerate(gantt.itertuples(index=False), 6):
            vals = [row.order_id, row.strategy_id, row.strategy_name, row.开始时间, row.结束时间, row.活跃小时, row.总出价, row.平均出价]
            for cidx, val in enumerate(vals, 1):
                ws.cell(ridx, cidx).value = val
            rows_h = gantt_h[(gantt_h["order_id"].eq(row.order_id)) & (gantt_h["strategy_id"].eq(row.strategy_id))]
            h_map = dict(zip(rows_h["hour"], rows_h["小时出价"]))
            for i, h in enumerate(hour_cols, 9):
                val = h_map.get(h, 0)
                if val:
                    cell = ws.cell(ridx, i)
                    cell.value = int(val)
                    cell.fill = PatternFill("solid", fgColor=BLUE_CELL)
                    cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

        write_sheet(writer, "APP竞价", "APP 竞价分析", "按 APP 出价量降序，展示竞价效率与出价价格。", app)
        write_sheet(writer, "时段竞价", "时段竞价分析", "按小时展示邀约、出价、竞得与平均出价。", hour)
        write_sheet(writer, "订单策略素材竞价", "订单 × 策略 × 素材竞价", "按出价量降序展示订单、策略、创意/素材的竞价效率。", order_material)
        write_sheet(writer, "出价分层", "出价分层分析", "按 bid_price 分层展示出价量、竞得率、点击率与花费。", bucket)
        for name, df in crosses.items():
            write_sheet(writer, name, name, "四个维度两两交叉分析：APP、时段、订单策略素材、出价分层。", df)
        write_sheet(writer, "口径说明", "口径说明", "数据源、字段关联和竞价专项分析规则。", notes)
        style_workbook(writer, total_spend)
        format_compare_block(writer.book["大盘竞价概览"])
        ws.freeze_panes = "I6"
        for col in range(9, 9 + len(hour_cols)):
            ws.column_dimensions[get_column_letter(col)].width = 8

    print(OUT.resolve())
    print("wide_rows", len(wide), "total_bid", int(total_bid), "sheets", 13)


if __name__ == "__main__":
    main()
