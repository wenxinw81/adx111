from __future__ import annotations

import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / ".vendor_pymysql"))
import pymysql  # noqa: E402


DAY = os.getenv("REPORT_DAY", "2026-07-25")
PREV_DAY = os.getenv(
    "REPORT_PREV_DAY",
    (datetime.strptime(DAY, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d"),
)
DAY_DT = datetime.strptime(DAY, "%Y-%m-%d")
OUT = Path(
    os.getenv(
        "REPORT_OUTPUT",
        f"outputs/adx_spend_report_{DAY_DT:%Y%m%d}/{DAY_DT.month}.{DAY_DT.day}广告投放花费分析报表.xlsx",
    )
)
DEFAULT_APP_MAPPING_CSV = Path(__file__).resolve().parents[1] / "data" / "app_mapping.csv"
APP_MAPPING_CSV = Path(os.getenv("APP_MAPPING_CSV", str(DEFAULT_APP_MAPPING_CSV)))
ORDER_ID_TEXT = os.getenv("REPORT_ORDER_ID", "").strip()
ORDER_ID = int(ORDER_ID_TEXT) if ORDER_ID_TEXT else None

BLUE_DARK = "073D63"
BLUE = "2F70C9"
BLUE_LIGHT = "DDEBF7"
BLUE_CELL = "40A2D8"
GREEN = "7FD0AA"
YELLOW = "FFE699"
RED = "FF0000"
GRID = "D9E2EC"
TEXT = "17324D"
WUHAN_TARGET_STRATEGY_IDS = {195, 199, 205, 212, 213}


def order_filter(alias: str = "r") -> str:
    return f" and {alias}.order_id = {ORDER_ID}" if ORDER_ID is not None else ""


def request_order_exists() -> str:
    if ORDER_ID is None:
        return ""
    return f"""
      and exists (
        select 1
        from ods.ods_ad_adx_bid_response_rt r
        where r.response_id=request_id
          and r.bid_imp_id=imp_id
          and r.order_id={ORDER_ID}
          and r.response_time >= request_time
          and r.response_time < date_add(request_time, interval 1 day)
      )
    """


def conn() -> pymysql.Connection:
    password = os.getenv("DORIS_PASSWORD")
    if not password:
        raise RuntimeError("DORIS_PASSWORD is required")
    return pymysql.connect(
        host=os.getenv("DORIS_HOST", "127.0.0.1"),
        port=int(os.getenv("DORIS_PORT", "19030")),
        user=os.getenv("DORIS_USER", "WishFox"),
        password=password,
        database=os.getenv("DORIS_DATABASE", "ads"),
        connect_timeout=10,
        read_timeout=int(os.getenv("DORIS_READ_TIMEOUT", "900")),
        write_timeout=int(os.getenv("DORIS_WRITE_TIMEOUT", "900")),
        charset="utf8mb4",
    )


def q(sql: str, params: tuple = ()) -> pd.DataFrame:
    attempts = int(os.getenv("DORIS_QUERY_RETRIES", "3"))
    for attempt in range(1, attempts + 1):
        try:
            with conn() as c:
                return pd.read_sql(sql, c, params=params)
        except pymysql.err.OperationalError as exc:
            code = exc.args[0] if exc.args else None
            if code not in {2006, 2013} or attempt >= attempts:
                raise
            wait_seconds = min(2**attempt, 8)
            print(f"Doris query lost connection ({code}); retry {attempt}/{attempts - 1} after {wait_seconds}s", file=sys.stderr)
            time.sleep(wait_seconds)
    raise RuntimeError("Doris query failed without returning a result")


def load_app_mapping() -> dict[str, str]:
    if not APP_MAPPING_CSV.exists():
        return {}
    df = pd.read_csv(APP_MAPPING_CSV, encoding="utf-8-sig")
    required = {"app_package_name", "app_name"}
    if not required.issubset(df.columns):
        return {}
    df = df[["app_package_name", "app_name", "is_verified", "request_count", "etl_time"]].copy()
    df["app_package_name"] = df["app_package_name"].fillna("").astype(str).str.strip()
    df["app_name"] = df["app_name"].fillna("").astype(str).str.strip()
    df = df[(df["app_package_name"].ne("")) & (df["app_name"].ne(""))]
    df["is_verified"] = pd.to_numeric(df["is_verified"], errors="coerce").fillna(0)
    df["request_count"] = pd.to_numeric(df["request_count"], errors="coerce").fillna(0)
    df = df.sort_values(["is_verified", "request_count", "etl_time"], ascending=[False, False, False])
    return df.drop_duplicates("app_package_name").set_index("app_package_name")["app_name"].to_dict()


def pct(num: pd.Series, den: pd.Series) -> pd.Series:
    return (num / den.replace(0, pd.NA)).fillna(0)


def spend_expr(alias: str = "") -> str:
    prefix = f"{alias}." if alias else ""
    return f"sum(coalesce(cast(nullif({prefix}wpr, '') as double), 0)) / 100000"


def add_spend_rates(df: pd.DataFrame, total_spend: float | None = None, has_invite: bool = True) -> pd.DataFrame:
    df = df.copy()
    if "邀约" in df and has_invite:
        df["邀约→出价"] = pct(df["出价"], df["邀约"])
    df["出价→竞得"] = pct(df["竞得"], df["出价"])
    df["竞得→点击"] = pct(df["点击"], df["竞得"])
    df["点击→落地页曝光"] = pct(df["落地页曝光"], df["点击"])
    df["落地页曝光→跳转"] = pct(df["跳转"], df["落地页曝光"])
    df["每次竞得花费"] = pct(df["花费(元)"], df["竞得"])
    df["每次点击花费"] = pct(df["花费(元)"], df["点击"])
    df["每次跳转花费"] = pct(df["花费(元)"], df["跳转"])
    if total_spend is not None:
        df["花费占比"] = df["花费(元)"] / total_spend if total_spend else 0
    return df


def adview_label(v) -> str:
    try:
        code = int(v)
    except (TypeError, ValueError):
        return "未知"
    names = {4: "插屏", 5: "开屏", 10: "信息流"}
    return f"{code}-{names.get(code, '其他')}"


def style_workbook(writer, total_spend: float) -> None:
    wb = writer.book
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"
        max_col = max(ws.max_column, 8)
        if ws["A1"].value:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws["A1"].fill = PatternFill("solid", fgColor=BLUE_DARK)
            ws["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 32
        if ws["A2"].value:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
            ws["A2"].fill = PatternFill("solid", fgColor="EAF2FB")
            ws["A2"].font = Font(name="Arial", size=10, color="5D7285")
            ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        side = Side(style="thin", color=GRID)
        border = Border(left=side, right=side, top=side, bottom=side)
        header_by_row: dict[int, int] = {}
        current_header = 4
        for row in range(4, ws.max_row + 1):
            labels = {str(ws.cell(row, c).value) for c in range(1, min(ws.max_column, 22) + 1) if ws.cell(row, c).value is not None}
            title_only = ws.cell(row, 1).value is not None and all(ws.cell(row, c).value is None for c in range(2, min(ws.max_column, 12) + 1))
            is_header = bool({"维度", "APP名称", "订单ID", "策略ID", "落地页ID", "hour", "日期", "指标", "order_id", "strategy_id", "地域", "广告位类型", "定向组", "总花费", "活跃小时"} & labels)
            if is_header:
                current_header = row
            header_by_row[row] = current_header
            if title_only and not is_header:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row, col)
                    cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
                    cell.font = Font(name="Arial", size=10, bold=True, color=TEXT)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = border
                continue
            if is_header or row == 4:
                for cell in ws[row]:
                    cell.fill = PatternFill("solid", fgColor=BLUE)
                    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
                ws.row_dimensions[row].height = 26
                continue
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.value is None:
                    continue
                header = str(ws.cell(header_by_row.get(row, 4), col).value or "")
                cell.font = Font(name="Arial", size=10, color=TEXT)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="right" if isinstance(cell.value, (int, float)) else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if isinstance(cell.value, float):
                    if "率" in header or "占比" in header or "→" in header:
                        cell.number_format = "0.00%"
                    else:
                        cell.number_format = "#,##0.00"
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
                if "花费" in header and isinstance(cell.value, (int, float)) and total_spend and cell.value / total_spend > 0.3:
                    cell.fill = PatternFill("solid", fgColor=YELLOW)
                if "跳转" in header and isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.font = Font(name="Arial", size=10, bold=True, color=RED)
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, min(ws.max_row, 250) + 1):
                val = ws.cell(row, col).value
                max_len = max(max_len, min(len(str(val)) if val is not None else 0, 50))
            ws.column_dimensions[get_column_letter(col)].width = max(10, min(max_len + 2, 38))


def write_sheet(writer, sheet_name: str, title: str, subtitle: str, df: pd.DataFrame) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
    ws = writer.book[sheet_name]
    ws["A1"] = title
    ws["A2"] = subtitle


def append_table(ws, start_row: int, title: str, df: pd.DataFrame) -> int:
    ws.cell(start_row, 1).value = title
    for cidx, col in enumerate(df.columns, 1):
        ws.cell(start_row + 1, cidx).value = col
    for ridx, row in enumerate(df.itertuples(index=False), start_row + 2):
        for cidx, val in enumerate(row, 1):
            ws.cell(ridx, cidx).value = val
    return start_row + len(df) + 4


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    app_map = load_app_mapping()

    daily_sql = f"""
    with
    req as (
      select date(request_time) d, count(*) invite
      from ods.ods_ad_adx_bid_request_rt
      where request_time >= %s and request_time < date_add(%s, interval 2 day)
      group by date(request_time)
    ),
    bid as (
      select date(response_time) d, count(*) bid, count(distinct strategy_id) strategy_cnt, count(distinct hour(response_time)) active_hours
      from ods.ods_ad_adx_bid_response_rt
      where response_time >= %s and response_time < date_add(%s, interval 2 day)
      group by date(response_time)
    ),
    win as (
      select date(response_time) d, count(*) win, {spend_expr()} spend
      from ods.ods_ad_adx_bid_response_result_rt
      where response_time >= %s and response_time < date_add(%s, interval 2 day)
      group by date(response_time)
    ),
    clk as (
      select date(response_time) d, count(*) click
      from ods.ods_ad_adx_bid_response_click_rt
      where response_time >= %s and response_time < date_add(%s, interval 2 day)
      group by date(response_time)
    ),
    clk_pairs as (
      select date(response_time) d, response_id, bid_seat_id
      from ods.ods_ad_adx_bid_response_click_rt
      where response_time >= %s and response_time < date_add(%s, interval 2 day)
      group by date(response_time), response_id, bid_seat_id
    ),
    track_keys as (
      select
        coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]rid=([^&]+)', 1), ''), t.root_id) response_id_key,
        coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]pyck=([^&]+)', 1), ''), t.parent_event_id) bid_seat_id_key,
        t.event_code
      from ods.ods_track_event_default_raw t
      where t.event_time >= %s and t.event_time < date_add(%s, interval 2 day)
    ),
    trk as (
      select c.d,
        count(distinct case when t.event_code='pageExp' then concat(t.response_id_key,'#',t.bid_seat_id_key) end) page_exp,
        count(distinct case when t.event_code='jumpButton' then concat(t.response_id_key,'#',t.bid_seat_id_key) end) jump
      from clk_pairs c
      join track_keys t on t.response_id_key=c.response_id and t.bid_seat_id_key=c.bid_seat_id
      group by c.d
    )
    select req.d 日期, invite 邀约, coalesce(bid,0) 出价, coalesce(win,0) 竞得, coalesce(click,0) 点击,
           coalesce(page_exp,0) 落地页曝光, coalesce(jump,0) 跳转, coalesce(spend,0) `花费(元)`,
           coalesce(active_hours,0) 活跃小时, coalesce(strategy_cnt,0) 投放策略数
    from req
    left join bid on req.d=bid.d
    left join win on req.d=win.d
    left join clk on req.d=clk.d
    left join trk on req.d=trk.d
    order by req.d
    """
    daily_params = (PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY)
    if ORDER_ID is not None:
        daily_sql = f"""
        with
        base as (
          select date(r.response_time) d, r.response_time, r.response_id, r.bid_seat_id, r.strategy_id
          from ods.ods_ad_adx_bid_response_rt r
          where r.response_time >= %s and r.response_time < date_add(%s, interval 2 day)
            and r.order_id = {ORDER_ID}
        ),
        req as (
          select d, count(*) invite
          from base
          group by d
        ),
        bid as (
          select d, count(*) bid, count(distinct strategy_id) strategy_cnt, count(distinct hour(response_time)) active_hours
          from base
          group by d
        ),
        win as (
          select b.d, count(*) win, sum(coalesce(cast(nullif(w.wpr, '') as double),0))/100000 spend
          from base b
          join ods.ods_ad_adx_bid_response_result_rt w
            on b.response_id=w.response_id and b.bid_seat_id=w.bid_seat_id
          where w.response_time >= %s and w.response_time < date_add(%s, interval 2 day)
          group by b.d
        ),
        clk_pairs as (
          select b.d, c.response_id, c.bid_seat_id
          from base b
          join ods.ods_ad_adx_bid_response_click_rt c
            on b.response_id=c.response_id and b.bid_seat_id=c.bid_seat_id
          where c.response_time >= %s and c.response_time < date_add(%s, interval 2 day)
          group by b.d, c.response_id, c.bid_seat_id
        ),
        clk as (
          select d, count(*) click
          from clk_pairs
          group by d
        ),
        track_keys as (
          select
            coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]rid=([^&]+)', 1), ''), t.root_id) response_id_key,
            coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]pyck=([^&]+)', 1), ''), t.parent_event_id) bid_seat_id_key,
            t.event_code
          from ods.ods_track_event_default_raw t
          where t.event_time >= %s and t.event_time < date_add(%s, interval 2 day)
        ),
        trk as (
          select c.d,
            count(distinct case when t.event_code='pageExp' then concat(t.response_id_key,'#',t.bid_seat_id_key) end) page_exp,
            count(distinct case when t.event_code='jumpButton' then concat(t.response_id_key,'#',t.bid_seat_id_key) end) jump
          from clk_pairs c
          join track_keys t on t.response_id_key=c.response_id and t.bid_seat_id_key=c.bid_seat_id
          group by c.d
        )
        select req.d 日期, invite 邀约, coalesce(bid,0) 出价, coalesce(win,0) 竞得, coalesce(click,0) 点击,
               coalesce(page_exp,0) 落地页曝光, coalesce(jump,0) 跳转, coalesce(spend,0) `花费(元)`,
               coalesce(active_hours,0) 活跃小时, coalesce(strategy_cnt,0) 投放策略数
        from req
        left join bid on req.d=bid.d
        left join win on req.d=win.d
        left join clk on req.d=clk.d
        left join trk on req.d=trk.d
        order by req.d
        """
        daily_params = (PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY, PREV_DAY)
    daily = q(daily_sql, daily_params)
    current = daily[daily["日期"].astype(str) == DAY].copy()
    prev = daily[daily["日期"].astype(str) == PREV_DAY].copy()
    if current.empty:
        raise RuntimeError(f"{DAY} 没有查询到数据")
    total_spend = float(current.iloc[0]["花费(元)"])

    wide_sql = f"""
    with
    win as (
      select response_id, bid_seat_id, count(*) win, sum(coalesce(cast(nullif(wpr, '') as double),0))/100000 spend_yuan
      from ods.ods_ad_adx_bid_response_result_rt
      where response_time >= %s and response_time < date_add(%s, interval 1 day)
      group by response_id, bid_seat_id
    ),
    clk as (
      select response_id, bid_seat_id, count(*) click
      from ods.ods_ad_adx_bid_response_click_rt
      where response_time >= %s and response_time < date_add(%s, interval 1 day)
      group by response_id, bid_seat_id
    ),
    track_keys as (
      select
        coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]rid=([^&]+)', 1), ''), t.root_id) response_id_key,
        coalesce(nullif(regexp_extract(cast(t.preset_attributes as string), '[?&]pyck=([^&]+)', 1), ''), t.parent_event_id) bid_seat_id_key,
        regexp_extract(cast(t.preset_attributes as string), '/dragIndex/([0-9]+)', 1) page_id,
        t.event_code
      from ods.ods_track_event_default_raw t
      where t.event_time >= %s and t.event_time < date_add(%s, interval 1 day)
    ),
    trk as (
      select t.response_id_key, t.bid_seat_id_key,
        max(t.page_id) page_id,
        max(case when t.event_code='pageExp' then 1 else 0 end) page_exp,
        max(case when t.event_code='pageViewComplete' then 1 else 0 end) complete,
        max(case when t.event_code in ('scrollStart','scrollEnd') then 1 else 0 end) scroll,
        max(case when t.event_code in ('vidPlay','vidPause') then 1 else 0 end) video,
        max(case when t.event_code='jumpButton' then 1 else 0 end) jump
      from track_keys t
      join clk c on t.response_id_key=c.response_id and t.bid_seat_id_key=c.bid_seat_id
      group by t.response_id_key, t.bid_seat_id_key
    ),
    dim_strategy as (
      select strategy_id, strategy_name, order_id, order_name
      from (
        select strategy_id, strategy_name, order_id, order_name,
               row_number() over(partition by strategy_id order by etl_time desc) rn
        from dim.dim_ad_dsp_strategy
      ) s
      where rn=1
    )
    select
      r.response_time,
      q.request_time,
      hour(q.request_time) hour,
      coalesce(nullif(q.app_bundle,''),'(空)') app_bundle,
      coalesce(nullif(q.geo_city,''),'地理缺失') geo_city,
      case
        when q.geo_lat is null or q.geo_lon is null then '地理缺失'
        when q.geo_lat between 29.8 and 31.6 and q.geo_lon between 113.6 and 115.2 then '武汉'
        else '非武汉'
      end geo_region,
      q.geo_lat,
      q.geo_lon,
      q.imp_adviewtype,
      q.imp_adlocation,
      q.imp_templateid,
      r.response_id,
      r.bid_seat_id,
      r.order_id,
      r.strategy_id,
      r.creative_id,
      coalesce(s.order_name, d.order_name, cast(r.order_id as string)) order_name,
      coalesce(s.strategy_name, d.strategy_name, cast(r.strategy_id as string)) strategy_name,
      coalesce(d.creative_name, cast(r.creative_id as string)) creative_name,
      d.material_id,
      d.material_type,
      r.land_page,
      r.curl,
      coalesce(win.win,0) win,
      coalesce(win.spend_yuan,0) spend_yuan,
      coalesce(clk.click,0) click,
      coalesce(trk.page_id,'') page_id,
      coalesce(trk.page_exp,0) page_exp,
      coalesce(trk.complete,0) complete,
      coalesce(trk.scroll,0) scroll,
      coalesce(trk.video,0) video,
      coalesce(trk.jump,0) jump
    from ods.ods_ad_adx_bid_response_rt r
    left join ods.ods_ad_adx_bid_request_rt q
      on q.request_id=r.response_id and q.imp_id=r.bid_imp_id
      and q.request_time >= %s and q.request_time < date_add(%s, interval 1 day)
    left join win on r.response_id=win.response_id and r.bid_seat_id=win.bid_seat_id
    left join clk on r.response_id=clk.response_id and r.bid_seat_id=clk.bid_seat_id
    left join trk on r.response_id=trk.response_id_key and r.bid_seat_id=trk.bid_seat_id_key
    left join dim_strategy s on r.strategy_id=s.strategy_id
    left join dim.dim_ad_dsp_creative d on trim(r.creative_id)=cast(d.creative_id as string)
    where r.response_time >= %s and r.response_time < date_add(%s, interval 1 day)
    {order_filter("r")}
    """
    wide = q(wide_sql, (DAY, DAY, DAY, DAY, DAY, DAY, DAY, DAY, DAY, DAY))
    wide["APP名称"] = wide["app_bundle"].map(app_map).fillna(wide["app_bundle"])
    wide["广告位类型"] = wide["imp_adviewtype"].map(adview_label)
    wide["定向组"] = wide["strategy_id"].apply(lambda x: "武汉定向" if pd.notna(x) and int(x) in WUHAN_TARGET_STRATEGY_IDS else "全国定向")
    wide["订单策略素材"] = (
        wide["order_name"].fillna("").astype(str)
        + " / "
        + wide["strategy_name"].fillna("").astype(str)
        + " / "
        + wide["creative_name"].fillna("").astype(str)
    )

    def bid_agg(keys: list[str], has_invite: bool = False) -> pd.DataFrame:
        g = wide.groupby(keys, dropna=False).agg(
            出价=("response_id", "size"),
            竞得=("win", "sum"),
            点击=("click", "sum"),
            落地页曝光=("page_exp", "sum"),
            跳转=("jump", "sum"),
            花费=("spend_yuan", "sum"),
            活跃小时=("hour", "nunique"),
        ).reset_index()
        g = g.rename(columns={"花费": "花费(元)"})
        return add_spend_rates(g, total_spend=total_spend, has_invite=has_invite)

    def request_agg(key_sql: str, label: str) -> pd.DataFrame:
        return q(
            f"""
            select {key_sql} {label}, count(*) 邀约
            from ods.ods_ad_adx_bid_request_rt
            where request_time >= %s and request_time < date_add(%s, interval 1 day)
            {request_order_exists()}
            group by {key_sql}
            """,
            (DAY, DAY),
        )

    app_req = request_agg("coalesce(nullif(app_bundle,''),'(空)')", "app_bundle")
    app_req["APP名称"] = app_req["app_bundle"].map(app_map).fillna(app_req["app_bundle"])
    app = pd.merge(app_req, bid_agg(["APP名称", "app_bundle"]), on=["APP名称", "app_bundle"], how="right").fillna(0)
    app = add_spend_rates(app, total_spend=total_spend).sort_values("花费(元)", ascending=False).head(50)
    app = app[["APP名称", "app_bundle"] + [c for c in app.columns if c not in {"APP名称", "app_bundle"}]]

    hour_req = request_agg("hour(request_time)", "hour")
    hour = pd.merge(hour_req, bid_agg(["hour"]), on="hour", how="right").fillna(0)
    hour = add_spend_rates(hour, total_spend=total_spend).sort_values("hour")

    adslot_req = request_agg("imp_adviewtype", "imp_adviewtype")
    adslot_req["广告位类型"] = adslot_req["imp_adviewtype"].map(adview_label)
    adslot_req = adslot_req.groupby("广告位类型", as_index=False)["邀约"].sum()
    adslot = pd.merge(adslot_req, bid_agg(["广告位类型"]), on="广告位类型", how="right").fillna(0)
    adslot = add_spend_rates(adslot, total_spend=total_spend).sort_values("花费(元)", ascending=False)

    geo_expr = """
      case
        when geo_lat is null or geo_lon is null then '地理缺失'
        when geo_lat between 29.8 and 31.6 and geo_lon between 113.6 and 115.2 then '武汉'
        else '非武汉'
      end
    """
    region_req = request_agg(geo_expr, "geo_region")
    region = pd.merge(region_req, bid_agg(["geo_region"]), on="geo_region", how="right").fillna(0)
    region = add_spend_rates(region, total_spend=total_spend).sort_values("花费(元)", ascending=False).head(80)
    region = region.rename(columns={"geo_region": "地域"})

    target_region = bid_agg(["定向组"])
    target_region.insert(1, "机会邀约", target_region["出价"])
    target_region = target_region[
        [
            "定向组",
            "花费(元)",
            "机会邀约",
            "出价",
            "竞得",
            "点击",
            "落地页曝光",
            "跳转",
            "每次竞得花费",
            "每次点击花费",
            "每次跳转花费",
            "出价→竞得",
            "竞得→点击",
            "点击→落地页曝光",
            "落地页曝光→跳转",
            "花费占比",
        ]
    ].sort_values("花费(元)", ascending=False)

    order_strategy = bid_agg(["order_id", "strategy_id", "creative_id", "order_name", "strategy_name", "creative_name", "material_id", "material_type", "订单策略素材"])
    order_strategy = order_strategy.sort_values("花费(元)", ascending=False).head(300)

    order_summary = bid_agg(["order_id", "order_name"]).sort_values("花费(元)", ascending=False)
    order_summary.insert(2, "机会邀约", order_summary["出价"])
    order_summary = order_summary[
        [
            "order_id",
            "order_name",
            "活跃小时",
            "花费(元)",
            "机会邀约",
            "出价",
            "竞得",
            "点击",
            "落地页曝光",
            "跳转",
            "每次竞得花费",
            "每次点击花费",
            "每次跳转花费",
            "出价→竞得",
            "竞得→点击",
            "点击→落地页曝光",
            "落地页曝光→跳转",
            "花费占比",
        ]
    ]

    strategy_summary = bid_agg(["order_id", "strategy_id", "order_name", "strategy_name"]).sort_values("花费(元)", ascending=False)
    strategy_summary.insert(4, "机会邀约", strategy_summary["出价"])
    strategy_summary = strategy_summary[
        [
            "order_id",
            "strategy_id",
            "order_name",
            "strategy_name",
            "活跃小时",
            "花费(元)",
            "机会邀约",
            "出价",
            "竞得",
            "点击",
            "落地页曝光",
            "跳转",
            "每次竞得花费",
            "每次点击花费",
            "每次跳转花费",
            "出价→竞得",
            "竞得→点击",
            "点击→落地页曝光",
            "落地页曝光→跳转",
            "花费占比",
        ]
    ]

    adslot_strategy = bid_agg(["strategy_id", "strategy_name", "广告位类型"])
    adslot_strategy = adslot_strategy[
        ["strategy_id", "strategy_name", "广告位类型", "花费(元)", "出价", "竞得", "点击", "每次竞得花费", "每次点击花费", "花费占比"]
    ].sort_values("花费(元)", ascending=False)

    landing = wide.groupby("page_id", dropna=False).agg(
        花费=("spend_yuan", "sum"),
        点击=("click", "sum"),
        落地页曝光=("page_exp", "sum"),
        跳转=("jump", "sum"),
        完读访问=("complete", "sum"),
        滚动访问=("scroll", "sum"),
        视频访问=("video", "sum"),
    ).reset_index().rename(columns={"page_id": "落地页ID", "花费": "花费(元)"})
    landing = landing[(landing["落地页ID"].ne("")) | (landing["落地页曝光"].gt(0)) | (landing["跳转"].gt(0))]
    landing["每次曝光花费"] = pct(landing["花费(元)"], landing["落地页曝光"])
    landing["每次跳转花费"] = pct(landing["花费(元)"], landing["跳转"])
    landing["点击→落地页曝光"] = pct(landing["落地页曝光"], landing["点击"])
    landing["落地页曝光→跳转"] = pct(landing["跳转"], landing["落地页曝光"])
    landing["完读率"] = pct(landing["完读访问"], landing["落地页曝光"])
    landing["花费占比"] = landing["花费(元)"] / total_spend if total_spend else 0
    landing = landing.sort_values("花费(元)", ascending=False)

    gantt_source = wide.groupby(["order_id", "strategy_id", "strategy_name", "hour"], dropna=False).agg(
        小时花费=("spend_yuan", "sum"),
        小时出价=("response_id", "size"),
    ).reset_index()
    gantt_base = wide.groupby(["order_id", "strategy_id", "strategy_name"], dropna=False).agg(
        开始时间=("response_time", "min"),
        结束时间=("response_time", "max"),
        活跃小时=("hour", "nunique"),
        总花费=("spend_yuan", "sum"),
        出价=("response_id", "size"),
        竞得=("win", "sum"),
    ).reset_index().rename(columns={"总花费": "总花费(元)"})
    hour_cols = sorted(int(x) for x in wide["hour"].dropna().unique())
    for h in hour_cols:
        spend_h = gantt_source[gantt_source["hour"].eq(h)][["order_id", "strategy_id", "小时花费", "小时出价"]].rename(
            columns={"小时花费": f"{h:02d}花费", "小时出价": f"{h:02d}出价"}
        )
        gantt_base = gantt_base.merge(spend_h, on=["order_id", "strategy_id"], how="left")
    gantt = gantt_base.sort_values("总花费(元)", ascending=False)

    dashboard = add_spend_rates(current.copy(), total_spend=total_spend)
    if not prev.empty:
        p = prev.iloc[0]
        c = current.iloc[0]
        compare = pd.DataFrame(
            [
                {"指标": m, DAY: c[m], PREV_DAY: p[m], "较前日变化": c[m] - p[m], "较前日变化率": (c[m] - p[m]) / p[m] if p[m] else 0}
                for m in ["花费(元)", "邀约", "出价", "竞得", "点击", "落地页曝光", "跳转"]
            ]
        )
    else:
        compare = pd.DataFrame()

    notes = pd.DataFrame(
        {
            "项目": ["花费口径", "漏斗定义", "订单限定", "埋点关联", "地域口径", "广告位类型", "策略名称"],
            "说明": [
                "花费=sum(wpr)/100000，单位元；按竞得结果表归因。",
                "邀约、出价、竞得、点击、落地页曝光(pageExp)、跳转(jumpButton)。",
                f"当前仅分析订单 {ORDER_ID}；请求表无订单字段，邀约统计为可关联到该订单出价的请求机会。" if ORDER_ID is not None else "未限定订单，按全量投放统计。",
                "track 优先从 preset_attributes.url 解析 rid/pyck 关联点击；解析不到时用 root_id/parent_event_id 兜底。",
                "地域优先使用经纬度矩形：武汉=纬度29.8–31.6且经度113.6–115.2；经纬度为空归为地理缺失。",
                "广告位类型使用 request.imp_adviewtype：4=插屏，5=开屏，10=信息流。",
                "strategy_id 名称优先来自 dim.dim_ad_dsp_strategy，创意/素材字段来自 dim.dim_ad_dsp_creative。",
            ],
        }
    )

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="大盘花费概览", index=False)
        ws = writer.book["大盘花费概览"]
        ws.delete_rows(1, ws.max_row)
        ws["A1"] = "广告投放花费专项分析"
        ws["A2"] = f"数据范围：{DAY} 00:00:00 至 {DAY} 23:59:59｜活跃小时：{int(current.iloc[0]['活跃小时'])}｜共投策略：{int(current.iloc[0]['投放策略数'])}｜花费：Σ(wpr÷100000)，单位：元"
        ws["A4"] = "大盘指标"
        metric_positions = [
            (5, 1, "总花费", "花费(元)"),
            (5, 2, "总邀约", "邀约"),
            (5, 3, "总出价", "出价"),
            (5, 4, "总竞得", "竞得"),
            (5, 5, "总点击", "点击"),
            (5, 6, "总跳转", "跳转"),
        ]
        cur_row = dashboard.iloc[0]
        for r, cidx, label, key in metric_positions:
            ws.cell(r, cidx).value = label
            ws.cell(r + 1, cidx).value = float(cur_row[key]) if key == "花费(元)" else int(cur_row[key])
        compact_metrics = [
            ("活跃小时", int(cur_row["活跃小时"])),
            ("花费/活跃小时", float(cur_row["花费(元)"] / cur_row["活跃小时"]) if cur_row["活跃小时"] else 0),
            ("单次竞得花费", float(cur_row["每次竞得花费"])),
            ("单次点击花费", float(cur_row["每次点击花费"])),
            ("单次跳转花费", float(cur_row["每次跳转花费"])),
            ("落地页到达率", float(cur_row["点击→落地页曝光"])),
        ]
        for cidx, (label, value) in enumerate(compact_metrics, 1):
            ws.cell(9, cidx).value = label
            ws.cell(10, cidx).value = value
        if not compare.empty:
            ws["A12"] = f"相较前一天（{PREV_DAY}）"
            next_row = append_table(ws, 13, "前日对比", compare)
        else:
            next_row = 12
        next_row = append_table(ws, next_row, "按订单", order_summary)
        append_table(ws, next_row, "按订单 × 策略 按花费降序", strategy_summary)
        write_sheet(writer, "APP花费", "APP 花费分析", "按花费降序；APP名称来自包名映射表，保留 app_bundle 用于核对。", app)
        write_sheet(writer, "时段花费", "时段花费分析", "按小时展示花费、漏斗、比率和单次成本；只展示有出价的小时。", hour)
        write_sheet(writer, "订单策略素材花费", "订单 × 策略 × 素材花费", "按花费降序展示订单、策略、创意/素材的花销效率。", order_strategy)
        pd.DataFrame().to_excel(writer, sheet_name="地域花费", index=False)
        ws_region = writer.book["地域花费"]
        ws_region.delete_rows(1, ws_region.max_row)
        ws_region["A1"] = "地域与定向花费对比"
        ws_region["A2"] = "定向组按策略配置划分：195、199、205、212、213=武汉定向，其他=全国定向；经纬度武汉口径为纬度 29.8–31.6、经度 113.6–115.2。"
        row_after_region = append_table(ws_region, 4, "按策略配置：武汉定向 vs 全国定向", target_region.rename(columns={"定向组": "维度"}))
        append_table(ws_region, row_after_region, "按请求经纬度：武汉 vs 非武汉", region.rename(columns={"地域": "维度"}))
        pd.DataFrame().to_excel(writer, sheet_name="广告位类型花费", index=False)
        ws_slot = writer.book["广告位类型花费"]
        ws_slot.delete_rows(1, ws_slot.max_row)
        ws_slot["A1"] = "广告位类型与花费"
        ws_slot["A2"] = "广告位取 request.imp_adviewtype：4=插屏，5=开屏，10=信息流。"
        row_after_slot = append_table(ws_slot, 4, "按广告位类型汇总", adslot.rename(columns={"广告位类型": "维度"}))
        append_table(ws_slot, row_after_slot, "策略与广告位关系", adslot_strategy.rename(columns={"strategy_id": "策略ID", "strategy_name": "策略名称"}))
        write_sheet(writer, "落地页花费", "落地页花费分析", "按落地页ID汇总花费、到达、跳转和页面互动。", landing)
        write_sheet(writer, "口径说明", "口径说明", "数据源、字段关联和花费专项分析规则。", notes)

        wb = writer.book
        ws = wb.create_sheet("策略花费甘特图", 1)
        total_cols = 8 + len(hour_cols)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws["A1"] = "策略小时花费甘特图"
        ws["A2"] = "蓝色单元格=该策略该小时有出价；数字=该小时花费(元)，有出价但未竞得显示0.00；仅展示单日活跃小时"
        fixed = ["订单ID", "策略ID", "策略名称", "开始时间", "结束时间", "活跃小时", "总花费(元)", "出价"]
        for cidx, h in enumerate(fixed, 1):
            ws.cell(4, cidx).value = h
        for i, h in enumerate(hour_cols, 9):
            ws.cell(4, i).value = f"{h:02d}"
        ws.cell(5, 1).value = "逐小时总花费"
        ws.cell(5, 7).value = total_spend
        hourly_spend = wide.groupby("hour")["spend_yuan"].sum().to_dict()
        for i, h in enumerate(hour_cols, 9):
            ws.cell(5, i).value = float(hourly_spend.get(h, 0))
        for ridx, (_, row) in enumerate(gantt.iterrows(), 6):
            vals = [row["order_id"], row["strategy_id"], row["strategy_name"], row["开始时间"], row["结束时间"], row["活跃小时"], row["总花费(元)"], row["出价"]]
            for cidx, val in enumerate(vals, 1):
                ws.cell(ridx, cidx).value = val
            for i, h in enumerate(hour_cols, 9):
                bid = row.get(f"{h:02d}出价")
                spend = row.get(f"{h:02d}花费")
                cell = ws.cell(ridx, i)
                if pd.notna(bid) and bid:
                    cell.value = float(spend or 0)
                    cell.fill = PatternFill("solid", fgColor=BLUE_CELL)
                    cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                    cell.number_format = "#,##0.00"
        style_workbook(writer, total_spend)
        ws.freeze_panes = "I6"
        for col in range(9, 9 + len(hour_cols)):
            ws.column_dimensions[get_column_letter(col)].width = 8

    print(OUT.resolve())
    print("wide_rows", len(wide), "total_spend", round(total_spend, 4), "sheets", 9)


if __name__ == "__main__":
    main()
