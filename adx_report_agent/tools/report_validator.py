from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..state import ValidationResult


EXPECTED_SHEETS = {
    "basic": [
        "大盘指标汇总",
        "过滤情况",
        "投放策略甘特图",
        "订单策略创意",
        "APP漏斗Top30",
        "落地页情况",
        "小时分析",
        "APP×时段",
        "APP×订单策略素材",
        "时段×订单策略素材",
        "口径说明",
    ],
    "spend": [
        "大盘花费概览",
        "策略花费甘特图",
        "APP花费",
        "时段花费",
        "订单策略素材花费",
        "地域花费",
        "广告位类型花费",
        "落地页花费",
        "口径说明",
    ],
    "bidding": [
        "大盘竞价概览",
        "投放策略甘特图",
        "APP竞价",
        "时段竞价",
        "订单策略素材竞价",
        "出价分层",
        "APP×时段",
        "APP×订单策略素材",
        "APP×出价分层",
        "时段×订单策略素材",
        "时段×出价分层",
        "订单策略素材×出价分层",
        "口径说明",
    ],
}


def _assert_app_top30(wb, errors: list[str], checks: list[str]) -> None:
    if "APP竞价" not in wb.sheetnames:
        return
    ws = wb["APP竞价"]
    row_count = max(ws.max_row - 4, 0)
    checks.append(f"APP竞价 rows={row_count}")
    if row_count > 30:
        errors.append(f"APP竞价超过Top30：当前 {row_count} 行")
    top_bundles = {str(ws.cell(r, 2).value) for r in range(5, ws.max_row + 1) if ws.cell(r, 2).value is not None}
    for sheet in ["APP×时段", "APP×订单策略素材", "APP×出价分层"]:
        if sheet not in wb.sheetnames:
            continue
        ws_cross = wb[sheet]
        bundles = {str(ws_cross.cell(r, 2).value) for r in range(5, ws_cross.max_row + 1) if ws_cross.cell(r, 2).value is not None}
        outside = bundles - top_bundles
        checks.append(f"{sheet} app_bundles={len(bundles)}")
        if outside:
            errors.append(f"{sheet} 含Top30外APP：{sorted(outside)[:5]}")


def validate_report_artifact(output_path: str | Path, analysis_type: str) -> ValidationResult:
    """Validate workbook shape and important business rules."""

    path = Path(output_path)
    checks: list[str] = []
    errors: list[str] = []
    if not path.exists():
        return ValidationResult(ok=False, errors=[f"输出文件不存在：{path}"])

    wb = load_workbook(path, data_only=True)
    expected = EXPECTED_SHEETS.get(analysis_type, [])
    missing = [sheet for sheet in expected if sheet not in wb.sheetnames]
    if missing:
        errors.append(f"缺少sheet：{missing}")
    checks.append(f"sheets={len(wb.sheetnames)}")

    for sheet in expected:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.max_row < 1:
                errors.append(f"{sheet} 工作表为空")
            checks.append(f"{sheet} rows={ws.max_row} cols={ws.max_column}")

    if analysis_type == "bidding":
        _assert_app_top30(wb, errors, checks)

    return ValidationResult(ok=not errors, checks=checks, errors=errors)
